import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import random
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Ensure output directory exists
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

# Read CSV files from the 'data' folder
courses_df = pd.read_csv("data/courses.csv")
config_df = pd.read_csv("data/config.csv").set_index("parameter")["value"]
rooms_df = pd.read_csv("data/rooms.csv")
sections_df = pd.read_csv("data/sections.csv")
faculty_df = pd.read_csv("data/faculty.csv")
assistants_df = pd.read_csv("data/assistants.csv")
elective_enrollments_df = pd.read_csv("data/elective_enrollments.csv").dropna()

# Extract configuration parameters
slot_duration = int(config_df["slot_duration_minutes"])  # 30 minutes
scheduling_days = config_df["scheduling_days"].split(";")
ta_threshold = int(config_df["teaching_assistant_threshold"])

# Define time slots (30-minute increments from 9:00 to 19:30)
start_time = datetime.strptime("09:00", "%H:%M")
end_time = datetime.strptime("19:30", "%H:%M")
time_slots = []
current_time = start_time
while current_time.time() <= end_time.time():
    time_slots.append(current_time.time().strftime("%H:%M"))
    current_time += timedelta(minutes=slot_duration)

# Define display slots (up to 19:30)
display_slots = [
    "09:00-09:30", "09:30-10:00", "10:00-10:30", "10:30-11:00", "11:00-11:30", "11:30-12:00",
    "12:00-12:30", "12:30-13:00", "13:00-13:30", "14:30-15:00", "15:00-15:30", "15:30-16:00",
    "16:00-16:30", "16:30-17:00", "17:00-17:30", "17:30-18:00", "18:00-18:30", "18:30-19:00",
    "19:00-19:30"
]

# Define breaks
morning_break_start = datetime.strptime("10:30", "%H:%M").time()
morning_break_end = datetime.strptime("11:00", "%H:%M").time()
lunch_schedule = {
    "CSE": {"start": datetime.strptime("13:00", "%H:%M").time(), "end": datetime.strptime("14:30", "%H:%M").time()},
    "DSAI": {"start": datetime.strptime("13:15", "%H:%M").time(), "end": datetime.strptime("14:45", "%H:%M").time()},
    "ECE": {"start": datetime.strptime("13:30", "%H:%M").time(), "end": datetime.strptime("15:00", "%H:%M").time()}
}

# Get unique semesters by department
semesters_by_dept = {
    "CSE": sorted(courses_df[courses_df["department"] == "CSE"]["semester"].unique()),
    "DSAI": sorted(courses_df[courses_df["department"] == "DSAI"]["semester"].unique()),
    "ECE": sorted(courses_df[courses_df["department"] == "ECE"]["semester"].unique()),
}

# Create a list of all timetable keys (dept_semester_section)
timetable_keys = []
for dept, semesters in semesters_by_dept.items():
    for semester in semesters:
        semester_courses = courses_df[(courses_df["department"] == dept) & (courses_df["semester"] == semester)]
        for section_id in semester_courses["section_id"].unique():
            key = f"{dept}_{semester}_{section_id}"
            timetable_keys.append(key)

# Initialize 3D schedule array: schedule[day][time_slot][timetable_key]
schedule = {
    day: {
        slot: {key: {} for key in timetable_keys}
        for slot in time_slots
    }
    for day in scheduling_days
}

# Track room usage to avoid conflicts: room_schedule[day][time_slot][room_number]
room_schedule = {
    day: {
        slot: {}
        for slot in time_slots
    }
    for day in scheduling_days
}

# Track slot availability for optimized randomization
slot_availability = {
    day: {slot: 0 for slot in time_slots}
    for day in scheduling_days
}

# Define colors for courses and baskets
color_palette = [
    "FFC1CC", "CCE5FF", "CCFFCC", "FFCC99", "E6CCFF", "FFFFCC",
    "FF9999", "99CCFF", "99FF99", "FFCC66", "CC99FF", "66CCCC",
    "FF66CC", "99CCCC", "FF9966", "CCCCFF"
]
course_colors = {}

def assign_color(identifier):
    if identifier not in course_colors:
        course_colors[identifier] = random.choice(color_palette)
    return course_colors[identifier]

# Helper functions
def get_faculty_name(faculty_ids):
    if pd.isna(faculty_ids):
        return "TBD"
    faculty_ids = str(faculty_ids).split(";")
    names = [faculty_df[faculty_df["faculty_id"] == int(fid)]["faculty_name"].iloc[0] for fid in faculty_ids if fid]
    return ", ".join(names)

def assign_room(enrollment, component_type, dept, course_code, day, start_slot, duration_slots):
    if component_type == "practical":
        room_type = "COMPUTER_LAB" if "CS" in course_code or "DS" in course_code else "HARDWARE_LAB"
        available_rooms = rooms_df[(rooms_df["type"] == room_type) & (rooms_df["capacity"] >= min(enrollment, 40))]
    else:
        available_rooms = rooms_df[rooms_df["type"].isin(["LECTURE_ROOM", "SEATER_120", "SEATER_240"]) & (rooms_df["capacity"] >= enrollment)]
    
    if available_rooms.empty:
        logging.warning(f"No available rooms for {course_code} ({component_type}) with enrollment {enrollment}")
        return None
    
    start_idx = time_slots.index(start_slot)
    slots_to_check = time_slots[start_idx:start_idx + duration_slots]
    
    for _, room in available_rooms.iterrows():
        room_number = room["room_number"]
        room_available = True
        for slot in slots_to_check:
            if room_schedule[day][slot].get(room_number):
                room_available = False
                break
        if room_available:
            for slot in slots_to_check:
                room_schedule[day][slot][room_number] = True
            return room_number
    logging.warning(f"No available room slots for {course_code} ({component_type}) on {day} at {start_slot}")
    return None

def is_slot_available(day, start_slot, duration_slots, timetable_key, faculty_ids, section_id, dept):
    start_time = datetime.strptime(start_slot, "%H:%M").time()
    end_time = (datetime.strptime(start_slot, "%H:%M") + timedelta(minutes=slot_duration * duration_slots)).time()
    lunch_start = lunch_schedule[dept]["start"]
    lunch_end = lunch_schedule[dept]["end"]
    if (start_time >= morning_break_start and start_time < morning_break_end) or \
       (end_time > morning_break_start and end_time <= morning_break_end) or \
       (start_time >= lunch_start and start_time < lunch_end) or \
       (end_time > lunch_start and end_time <= lunch_end):
        return False
    
    start_idx = time_slots.index(start_slot)
    if start_idx + duration_slots > len(time_slots):
        return False
    
    slots_to_check = time_slots[start_idx:start_idx + duration_slots]
    
    faculty_ids = set(str(faculty_ids).split(";")) if not pd.isna(faculty_ids) else set()
    for slot in slots_to_check:
        slot_schedule = schedule[day][slot]
        if slot_schedule[timetable_key]:
            return False
        for key in timetable_keys:
            if slot_schedule[key]:
                other_faculty = set(str(slot_schedule[key].get("faculty_ids", "")).split(";"))
                if faculty_ids & other_faculty:
                    return False
    return True

def update_slot_availability():
    for day in scheduling_days:
        for slot in time_slots:
            slot_usage = 0
            for key in timetable_keys:
                if schedule[day][slot].get(key):
                    slot_usage += 1
            slot_availability[day][slot] = slot_usage

def get_available_slots(day, duration_slots, timetable_key, faculty_ids, section_id, dept):
    available_slots = []
    for start_slot in time_slots[:-duration_slots + 1]:
        if is_slot_available(day, start_slot, duration_slots, timetable_key, faculty_ids, section_id, dept):
            available_slots.append(start_slot)
    available_slots.sort(key=lambda slot: slot_availability[day][slot])
    return available_slots

# Handle electives: group by basket
basket_courses = {}
for _, course in courses_df.iterrows():
    if course["is_elective"] and pd.notna(course["basket_id"]):
        basket_id = course["basket_id"]
        key = (course["department"], course["semester"], basket_id)
        if key not in basket_courses:
            basket_courses[key] = []
        basket_courses[key].append(course)

# Log the detected elective baskets
logging.info("Detected elective baskets:")
for key, courses in basket_courses.items():
    dept, semester, basket_id = key
    course_codes = [course["course_code"] for course in courses]
    logging.info(f" - {basket_id} in {dept} semester {semester}: {course_codes}")

# Process baskets (all courses in a basket have the same LTPSC)
basket_schedules = {}
for key, courses in basket_courses.items():
    dept, semester, basket_id = key
    # Use the LTPSC of the first course since all courses in the basket have the same LTPSC
    standard_ltpsc = (courses[0]["lecture_hours"], courses[0]["tutorial_hours"], courses[0]["practical_hours"], courses[0]["self_study_hours"], courses[0]["credits"])
    basket_schedules[key] = {"ltpsc": standard_ltpsc, "courses": []}
    for course in courses:
        basket_schedules[key]["courses"].append({
            "course": course,
            "timetable_key": f"{course['department']}_{course['semester']}_{course['section_id']}",
            "enrollment": elective_enrollments_df[
                (elective_enrollments_df["course_id"] == course["course_id"]) &
                (elective_enrollments_df["section_id"] == course["section_id"])
            ]["enrollment"].iloc[0] if not elective_enrollments_df[
                (elective_enrollments_df["course_id"] == course["course_id"]) &
                (elective_enrollments_df["section_id"] == course["section_id"])
            ].empty else course["enrollment"]
        })

# Calculate total items for progress tracking
total_items = len(courses_df[courses_df["is_elective"] == False]) + len(basket_schedules)
items_processed = 0

# Group combined courses
combined_courses = {}
for _, course in courses_df.iterrows():
    if course["combined"]:
        course_key = (course["course_code"], course["faculty_ids"])
        if course_key not in combined_courses:
            combined_courses[course_key] = []
        combined_courses[course_key].append({
            "course": course,
            "timetable_key": f"{course['department']}_{course['semester']}_{course['section_id']}",
            "enrollment": course["enrollment"]
        })

# Store elective scheduling details for output
elective_details = []

# Schedule combined courses first
max_attempts = 100
logging.info("Starting to schedule combined courses")
for (course_code, faculty_ids), instances in combined_courses.items():
    total_enrollment = sum(instance["enrollment"] for instance in instances)
    course = instances[0]["course"]
    timetable_keys_for_course = [instance["timetable_key"] for instance in instances]
    dept = course["department"]
    
    logging.info(f"Scheduling combined course {course_code} for {len(timetable_keys_for_course)} sections")
    
    lecture_slots = 3 if course["lecture_hours"] >= 1.5 else 2
    lecture_sessions = 2 if course["lecture_hours"] == 3 else 1
    tutorial_slots = 2 if course["tutorial_hours"] > 0 else 0
    practical_slots = 4 if course["practical_hours"] > 0 else 0
    
    # Schedule practicals first
    if practical_slots > 0:
        lab_capacity = rooms_df[rooms_df["type"].isin(["COMPUTER_LAB", "HARDWARE_LAB"])]["capacity"].min()
        batches = max(1, int(total_enrollment / lab_capacity) + (1 if total_enrollment % lab_capacity else 0))
        for batch in range(batches):
            attempts = 0
            scheduled = False
            days_shuffled = scheduling_days.copy()
            random.shuffle(days_shuffled)
            for day in days_shuffled:
                update_slot_availability()
                available_slots = get_available_slots(day, practical_slots, timetable_keys_for_course[0], faculty_ids, course["section_id"], dept)
                for start_slot in available_slots:
                    attempts += 1
                    logging.info(f"Attempt {attempts} to schedule practical (Batch {chr(65+batch)}) for {course_code} on {day} at {start_slot}")
                    room = assign_room(min(total_enrollment, lab_capacity), "practical", dept, course_code, day, start_slot, practical_slots)
                    if room:
                        start_idx = time_slots.index(start_slot)
                        for idx in range(start_idx, start_idx + practical_slots):
                            slot = time_slots[idx]
                            for key in timetable_keys_for_course:
                                schedule[day][slot][key] = {
                                    "label": f"{course_code} (LAB) (Batch {chr(65+batch)})\n{room}",
                                    "course_code": course_code,
                                    "faculty_ids": faculty_ids,
                                    "section_id": course["section_id"],
                                    "component": "practical"
                                }
                        logging.info(f"Successfully scheduled practical (Batch {chr(65+batch)}) for {course_code} on {day} at {start_slot}")
                        scheduled = True
                        break
                if scheduled:
                    break
            if not scheduled:
                logging.warning(f"Failed to schedule practical (Batch {chr(65+batch)}) for combined course {course_code} after {attempts} attempts")
    
    # Schedule lecture sessions
    lecture_days = []
    for session in range(lecture_sessions):
        attempts = 0
        scheduled = False
        days_shuffled = scheduling_days.copy()
        random.shuffle(days_shuffled)
        for day in days_shuffled:
            if day in lecture_days:
                continue
            update_slot_availability()
            available_slots = get_available_slots(day, lecture_slots, timetable_keys_for_course[0], faculty_ids, course["section_id"], dept)
            for start_slot in available_slots:
                attempts += 1
                logging.info(f"Attempt {attempts} to schedule lecture session {session+1} for {course_code} on {day} at {start_slot}")
                room = assign_room(total_enrollment, "lecture", dept, course_code, day, start_slot, lecture_slots)
                if room:
                    start_idx = time_slots.index(start_slot)
                    for idx in range(start_idx, start_idx + lecture_slots):
                        slot = time_slots[idx]
                        for key in timetable_keys_for_course:
                            schedule[day][slot][key] = {
                                "label": f"{course_code} (L)\n{room}",
                                "course_code": course_code,
                                "faculty_ids": faculty_ids,
                                "section_id": course["section_id"],
                                "component": "lecture"
                            }
                    lecture_days.append(day)
                    logging.info(f"Successfully scheduled lecture session {session+1} for {course_code} on {day} at {start_slot}")
                    scheduled = True
                    break
            if scheduled:
                break
        if not scheduled:
            logging.warning(f"Failed to schedule lecture session {session+1} for combined course {course_code} after {attempts} attempts")
    
    # Schedule tutorial
    if tutorial_slots > 0:
        attempts = 0
        scheduled = False
        days_shuffled = scheduling_days.copy()
        random.shuffle(days_shuffled)
        for day in days_shuffled:
            if day in lecture_days:
                continue
            update_slot_availability()
            available_slots = get_available_slots(day, tutorial_slots, timetable_keys_for_course[0], faculty_ids, course["section_id"], dept)
            for start_slot in available_slots:
                attempts += 1
                logging.info(f"Attempt {attempts} to schedule tutorial for {course_code} on {day} at {start_slot}")
                room = assign_room(total_enrollment, "tutorial", dept, course_code, day, start_slot, tutorial_slots)
                if room:
                    start_idx = time_slots.index(start_slot)
                    for idx in range(start_idx, start_idx + tutorial_slots):
                        slot = time_slots[idx]
                        for key in timetable_keys_for_course:
                            schedule[day][slot][key] = {
                                "label": f"{course_code} (T)\n{room}",
                                "course_code": course_code,
                                "faculty_ids": faculty_ids,
                                "section_id": course["section_id"],
                                "component": "tutorial"
                            }
                    logging.info(f"Successfully scheduled tutorial for {course_code} on {day} at {start_slot}")
                    scheduled = True
                    break
            if scheduled:
                break
        if not scheduled:
            logging.warning(f"Failed to schedule tutorial for combined course {course_code} after {attempts} attempts")
    
    items_processed += 1
    progress = (items_processed / total_items) * 100
    logging.info(f"Progress: {items_processed}/{total_items} items scheduled ({progress:.2f}%)")

# Schedule elective baskets (schedule only once per basket)
logging.info("Starting to schedule elective baskets")
for key, basket_data in basket_schedules.items():
    dept, semester, basket_id = key
    ltpsc = basket_data["ltpsc"]
    courses = basket_data["courses"]
    timetable_keys_for_basket = [course["timetable_key"] for course in courses]
    faculty_ids_set = set()
    for course_data in courses:
        course = course_data["course"]
        if not pd.isna(course["faculty_ids"]):
            faculty_ids_set.update(str(course["faculty_ids"]).split(";"))
    
    logging.info(f"Scheduling elective basket {basket_id} in {dept} semester {semester} with LTPSC {ltpsc}")
    
    lecture_slots = 3 if ltpsc[0] >= 1.5 else 2
    lecture_sessions = 2 if ltpsc[0] == 3 else 1
    tutorial_slots = 2 if ltpsc[1] > 0 else 0
    practical_slots = 4 if ltpsc[2] > 0 else 0
    
    # Schedule practicals first
    practical_time_slots = []
    if practical_slots > 0:
        lab_capacity = rooms_df[rooms_df["type"].isin(["COMPUTER_LAB", "HARDWARE_LAB"])]["capacity"].min()
        # Find max enrollment for batch calculation
        max_enrollment = max(course["enrollment"] for course in courses)
        batches = max(1, int(max_enrollment / lab_capacity) + (1 if max_enrollment % lab_capacity else 0))
        for batch in range(batches):
            attempts = 0
            scheduled = False
            days_shuffled = scheduling_days.copy()
            random.shuffle(days_shuffled)
            for day in days_shuffled:
                update_slot_availability()
                available_slots = get_available_slots(day, practical_slots, timetable_keys_for_basket[0], ";".join(faculty_ids_set), courses[0]["course"]["section_id"], dept)
                for start_slot in available_slots:
                    attempts += 1
                    logging.info(f"Attempt {attempts} to schedule practical (Batch {chr(65+batch)}) for basket {basket_id} on {day} at {start_slot}")
                    rooms = {}
                    for course_data in courses:
                        course = course_data["course"]
                        enrollment = course_data["enrollment"]
                        room = assign_room(min(enrollment, lab_capacity), "practical", dept, course["course_code"], day, start_slot, practical_slots)
                        if room:
                            rooms[course["course_code"]] = room
                        else:
                            for slot in time_slots[time_slots.index(start_slot):time_slots.index(start_slot) + practical_slots]:
                                for r in rooms.values():
                                    if room_schedule[day][slot].get(r):
                                        del room_schedule[day][slot][r]
                            break
                    if len(rooms) == len(courses):
                        start_idx = time_slots.index(start_slot)
                        time_slot_range = f"{start_slot}-{time_slots[start_idx + practical_slots - 1]}"
                        practical_time_slots.append((day, time_slot_range))
                        room_assignments = "\n".join([f"{course_code}: {room}" for course_code, room in rooms.items()])
                        for idx in range(start_idx, start_idx + practical_slots):
                            slot = time_slots[idx]
                            for key in timetable_keys_for_basket:
                                schedule[day][slot][key] = {
                                    "label": f"{basket_id} (LAB) (Batch {chr(65+batch)})\n{room_assignments}",
                                    "course_code": basket_id,
                                    "faculty_ids": ";".join(faculty_ids_set),
                                    "section_id": courses[0]["course"]["section_id"],
                                    "component": "practical"
                                }
                        for course_data in courses:
                            course = course_data["course"]
                            elective_details.append({
                                "Basket ID": basket_id,
                                "Course Name": course["course_code"],
                                "Faculty": get_faculty_name(course["faculty_ids"]),
                                "Room": rooms[course["course_code"]],
                                "Time Slot": f"LAB (Batch {chr(65+batch)}): {day} {time_slot_range}",
                                "LTPSC": f"{ltpsc[0]}-{ltpsc[1]}-{ltpsc[2]}-{ltpsc[3]}-{ltpsc[4]}",
                                "Extra Sessions": ""
                            })
                        logging.info(f"Successfully scheduled practical (Batch {chr(65+batch)}) for basket {basket_id} on {day} at {start_slot}")
                        scheduled = True
                        break
                if scheduled:
                    break
            if not scheduled:
                logging.warning(f"Failed to schedule practical (Batch {chr(65+batch)}) for basket {basket_id} after {attempts} attempts")
    
    # Schedule lecture sessions
    lecture_days = []
    lecture_time_slots = []
    for session in range(lecture_sessions):
        attempts = 0
        scheduled = False
        days_shuffled = scheduling_days.copy()
        random.shuffle(days_shuffled)
        for day in days_shuffled:
            if day in lecture_days:
                continue
            update_slot_availability()
            available_slots = get_available_slots(day, lecture_slots, timetable_keys_for_basket[0], ";".join(faculty_ids_set), courses[0]["course"]["section_id"], dept)
            for start_slot in available_slots:
                attempts += 1
                logging.info(f"Attempt {attempts} to schedule lecture session {session+1} for basket {basket_id} on {day} at {start_slot}")
                rooms = {}
                for course_data in courses:
                    course = course_data["course"]
                    enrollment = course_data["enrollment"]
                    room = assign_room(enrollment, "lecture", dept, course["course_code"], day, start_slot, lecture_slots)
                    if room:
                        rooms[course["course_code"]] = room
                    else:
                        for slot in time_slots[time_slots.index(start_slot):time_slots.index(start_slot) + lecture_slots]:
                            for r in rooms.values():
                                if room_schedule[day][slot].get(r):
                                    del room_schedule[day][slot][r]
                        break
                if len(rooms) == len(courses):
                    start_idx = time_slots.index(start_slot)
                    time_slot_range = f"{start_slot}-{time_slots[start_idx + lecture_slots - 1]}"
                    lecture_time_slots.append((day, time_slot_range))
                    room_assignments = "\n".join([f"{course_code}: {room}" for course_code, room in rooms.items()])
                    for idx in range(start_idx, start_idx + lecture_slots):
                        slot = time_slots[idx]
                        for key in timetable_keys_for_basket:
                            schedule[day][slot][key] = {
                                "label": f"{basket_id} (L)\n{room_assignments}",
                                "course_code": basket_id,
                                "faculty_ids": ";".join(faculty_ids_set),
                                "section_id": courses[0]["course"]["section_id"],
                                "component": "lecture"
                            }
                    for course_data in courses:
                        course = course_data["course"]
                        entry_exists = False
                        for detail in elective_details:
                            if detail["Course Name"] == course["course_code"] and detail["Basket ID"] == basket_id:
                                detail["Time Slot"] = f"L: {day} {time_slot_range}, " + detail["Time Slot"]
                                entry_exists = True
                                break
                        if not entry_exists:
                            elective_details.append({
                                "Basket ID": basket_id,
                                "Course Name": course["course_code"],
                                "Faculty": get_faculty_name(course["faculty_ids"]),
                                "Room": rooms[course["course_code"]],
                                "Time Slot": f"L: {day} {time_slot_range}",
                                "LTPSC": f"{ltpsc[0]}-{ltpsc[1]}-{ltpsc[2]}-{ltpsc[3]}-{ltpsc[4]}",
                                "Extra Sessions": ""
                            })
                    lecture_days.append(day)
                    logging.info(f"Successfully scheduled lecture session {session+1} for basket {basket_id} on {day} at {start_slot}")
                    scheduled = True
                    break
            if scheduled:
                break
        if not scheduled:
            logging.warning(f"Failed to schedule lecture session {session+1} for basket {basket_id} after {attempts} attempts")
    
    # Schedule tutorial
    tutorial_time_slots = []
    if tutorial_slots > 0:
        attempts = 0
        scheduled = False
        days_shuffled = scheduling_days.copy()
        random.shuffle(days_shuffled)
        for day in days_shuffled:
            if day in lecture_days:
                continue
            update_slot_availability()
            available_slots = get_available_slots(day, tutorial_slots, timetable_keys_for_basket[0], ";".join(faculty_ids_set), courses[0]["course"]["section_id"], dept)
            for start_slot in available_slots:
                attempts += 1
                logging.info(f"Attempt {attempts} to schedule tutorial for basket {basket_id} on {day} at {start_slot}")
                rooms = {}
                for course_data in courses:
                    course = course_data["course"]
                    enrollment = course_data["enrollment"]
                    room = assign_room(enrollment, "tutorial", dept, course["course_code"], day, start_slot, tutorial_slots)
                    if room:
                        rooms[course["course_code"]] = room
                    else:
                        for slot in time_slots[time_slots.index(start_slot):time_slots.index(start_slot) + tutorial_slots]:
                            for r in rooms.values():
                                if room_schedule[day][slot].get(r):
                                    del room_schedule[day][slot][r]
                        break
                if len(rooms) == len(courses):
                    start_idx = time_slots.index(start_slot)
                    time_slot_range = f"{start_slot}-{time_slots[start_idx + tutorial_slots - 1]}"
                    tutorial_time_slots.append((day, time_slot_range))
                    room_assignments = "\n".join([f"{course_code}: {room}" for course_code, room in rooms.items()])
                    for idx in range(start_idx, start_idx + tutorial_slots):
                        slot = time_slots[idx]
                        for key in timetable_keys_for_basket:
                            schedule[day][slot][key] = {
                                "label": f"{basket_id} (T)\n{room_assignments}",
                                "course_code": basket_id,
                                "faculty_ids": ";".join(faculty_ids_set),
                                "section_id": courses[0]["course"]["section_id"],
                                "component": "tutorial"
                            }
                    for detail in elective_details:
                        if detail["Basket ID"] == basket_id:
                            detail["Time Slot"] += f", T: {day} {time_slot_range}"
                    logging.info(f"Successfully scheduled tutorial for basket {basket_id} on {day} at {start_slot}")
                    scheduled = True
                    break
            if scheduled:
                break
        if not scheduled:
            logging.warning(f"Failed to schedule tutorial for basket {basket_id} after {attempts} attempts")
    
    items_processed += 1
    progress = (items_processed / total_items) * 100
    logging.info(f"Progress: {items_processed}/{total_items} items scheduled ({progress:.2f}%)")

# Schedule non-elective, non-combined courses
logging.info("Starting to schedule non-elective, non-combined courses")
for dept, semesters in semesters_by_dept.items():
    logging.info(f"Processing department: {dept}")
    for semester in semesters:
        logging.info(f"Processing semester: {semester}")
        semester_courses = courses_df[(courses_df["department"] == dept) & (courses_df["semester"] == semester)]
        for section_id in semester_courses["section_id"].unique():
            timetable_key = f"{dept}_{semester}_{section_id}"
            logging.info(f"Processing section: {timetable_key}")
            section_courses = semester_courses[semester_courses["section_id"] == section_id]
            for _, course in section_courses.iterrows():
                if course["combined"] or course["is_elective"]:
                    continue
                course_code = course["course_code"]
                logging.info(f"Scheduling course: {course_code}")
                enrollment = course["enrollment"]
                
                lecture_slots = 3 if course["lecture_hours"] >= 1.5 else 2
                lecture_sessions = 2 if course["lecture_hours"] == 3 else 1
                tutorial_slots = 2 if course["tutorial_hours"] > 0 else 0
                practical_slots = 4 if course["practical_hours"] > 0 else 0
                
                # Schedule practicals first
                if practical_slots > 0:
                    lab_capacity = rooms_df[rooms_df["type"].isin(["COMPUTER_LAB", "HARDWARE_LAB"])]["capacity"].min()
                    batches = max(1, int(enrollment / lab_capacity) + (1 if enrollment % lab_capacity else 0))
                    for batch in range(batches):
                        attempts = 0
                        scheduled = False
                        days_shuffled = scheduling_days.copy()
                        random.shuffle(days_shuffled)
                        for day in days_shuffled:
                            update_slot_availability()
                            available_slots = get_available_slots(day, practical_slots, timetable_key, course["faculty_ids"], section_id, dept)
                            for start_slot in available_slots:
                                attempts += 1
                                logging.info(f"Attempt {attempts} to schedule practical (Batch {chr(65+batch)}) for {course_code} on {day} at {start_slot}")
                                room = assign_room(min(enrollment, lab_capacity), "practical", dept, course["course_code"], day, start_slot, practical_slots)
                                if room:
                                    start_idx = time_slots.index(start_slot)
                                    for idx in range(start_idx, start_idx + practical_slots):
                                        slot = time_slots[idx]
                                        schedule[day][slot][timetable_key] = {
                                            "label": f"{course['course_code']} (LAB) (Batch {chr(65+batch)})\n{room}",
                                            "course_code": course["course_code"],
                                            "faculty_ids": course["faculty_ids"],
                                            "section_id": section_id,
                                            "component": "practical"
                                        }
                                    logging.info(f"Successfully scheduled practical (Batch {chr(65+batch)}) for {course_code} on {day} at {start_slot}")
                                    scheduled = True
                                    break
                            if scheduled:
                                break
                        if not scheduled:
                            logging.warning(f"Failed to schedule practical (Batch {chr(65+batch)}) for course {course_code} in {timetable_key} after {attempts} attempts")
                
                # Schedule lecture sessions
                lecture_days = []
                for session in range(lecture_sessions):
                    attempts = 0
                    scheduled = False
                    days_shuffled = scheduling_days.copy()
                    random.shuffle(days_shuffled)
                    for day in days_shuffled:
                        if day in lecture_days:
                            continue
                        update_slot_availability()
                        available_slots = get_available_slots(day, lecture_slots, timetable_key, course["faculty_ids"], section_id, dept)
                        for start_slot in available_slots:
                            attempts += 1
                            logging.info(f"Attempt {attempts} to schedule lecture session {session+1} for {course_code} on {day} at {start_slot}")
                            room = assign_room(enrollment, "lecture", dept, course["course_code"], day, start_slot, lecture_slots)
                            if room:
                                start_idx = time_slots.index(start_slot)
                                for idx in range(start_idx, start_idx + lecture_slots):
                                    slot = time_slots[idx]
                                    schedule[day][slot][timetable_key] = {
                                        "label": f"{course['course_code']} (L)\n{room}",
                                        "course_code": course["course_code"],
                                        "faculty_ids": course["faculty_ids"],
                                        "section_id": section_id,
                                        "component": "lecture"
                                    }
                                lecture_days.append(day)
                                logging.info(f"Successfully scheduled lecture session {session+1} for {course_code} on {day} at {start_slot}")
                                scheduled = True
                                break
                        if scheduled:
                            break
                    if not scheduled:
                        logging.warning(f"Failed to schedule lecture session {session+1} for course {course_code} in {timetable_key} after {attempts} attempts")
                
                # Schedule tutorial
                if tutorial_slots > 0:
                    attempts = 0
                    scheduled = False
                    days_shuffled = scheduling_days.copy()
                    random.shuffle(days_shuffled)
                    for day in days_shuffled:
                        if day in lecture_days:
                            continue
                        update_slot_availability()
                        available_slots = get_available_slots(day, tutorial_slots, timetable_key, course["faculty_ids"], section_id, dept)
                        for start_slot in available_slots:
                            attempts += 1
                            logging.info(f"Attempt {attempts} to schedule tutorial for {course_code} on {day} at {start_slot}")
                            room = assign_room(enrollment, "tutorial", dept, course["course_code"], day, start_slot, tutorial_slots)
                            if room:
                                start_idx = time_slots.index(start_slot)
                                for idx in range(start_idx, start_idx + tutorial_slots):
                                    slot = time_slots[idx]
                                    schedule[day][slot][timetable_key] = {
                                        "label": f"{course['course_code']} (T)\n{room}",
                                        "course_code": course["course_code"],
                                        "faculty_ids": course["faculty_ids"],
                                        "section_id": section_id,
                                        "component": "tutorial"
                                    }
                                logging.info(f"Successfully scheduled tutorial for {course_code} on {day} at {start_slot}")
                                scheduled = True
                                break
                        if scheduled:
                            break
                    if not scheduled:
                        logging.warning(f"Failed to schedule tutorial for course {course_code} in {timetable_key} after {attempts} attempts")
                
                items_processed += 1
                progress = (items_processed / total_items) * 100
                logging.info(f"Progress: {items_processed}/{total_items} items scheduled ({progress:.2f}%)")

# Map time slots to display slots
slot_mapping = {}
for display_slot in display_slots:
    start, end = display_slot.split("-")
    start_time = datetime.strptime(start, "%H:%M").time()
    end_time = datetime.strptime(end, "%H:%M").time()
    slot_mapping[display_slot] = [slot for slot in time_slots if start_time <= datetime.strptime(slot, "%H:%M").time() < end_time]

# Generate HTML with timetable and elective details
logging.info("Generating HTML output")
html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Timetable</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        .break-cell { background-color: #d3d3d3; text-align: center; }
        .timetable-table { width: 100%; border-collapse: collapse; margin-bottom: 2rem; }
        .timetable-table th, .timetable-table td { border: 1px solid #000; padding: 0.5rem; text-align: center; }
        .timetable-table th { background-color: #f0f0f0; }
        .elective-table { width: 100%; border-collapse: collapse; margin-top: 2rem; }
        .elective-table th, .elective-table td { border: 1px solid #000; padding: 0.5rem; text-align: left; }
        .elective-table th { background-color: #e0e0e0; }
    </style>
</head>
<body class="bg-gray-100 p-8">
    <h1 class="text-3xl font-bold mb-4 text-center">INDIAN INSTITUTE OF INFORMATION TECHNOLOGY, DHARWAD</h1>
    <h2 class="text-2xl font-semibold mb-8 text-center">Time Table for an Academic year Dec 24 – April 2025</h2>
"""
for dept, semesters in semesters_by_dept.items():
    for semester in semesters:
        semester_courses = courses_df[(courses_df["department"] == dept) & (courses_df["semester"] == semester)]
        for section_id in semester_courses["section_id"].unique():
            timetable_key = f"{dept}_{semester}_{section_id}"
            section = sections_df[sections_df["section_id"] == section_id].iloc[0]
            roll_start = f"{section['year']}{section['department'].lower()}{section['batch_name'][-2:]}001"
            roll_end = f"{section['year']}{section['department'].lower()}{section['batch_name'][-2:]}0{int(section['strength']):02d}"
            html_content += f'<h3 class="text-xl font-semibold mb-2">Section: {section["batch_name"]} – Roll no {roll_start} to {roll_end}</h3>'
            html_content += f'<p class="mb-4">Group mail id – {section["year"]}{section["department"].lower()}{section["batch_name"][-2:]}@iiitdwd.ac.in</p>'
            html_content += '<table class="timetable-table">'
            html_content += '<thead><tr><th>Day</th>'
            for slot in display_slots:
                html_content += f'<th>{slot}</th>'
            html_content += '</tr></thead><tbody>'
            
            for day in scheduling_days:
                html_content += f'<tr><td>{day}</td>'
                for display_slot in display_slots:
                    start_time = datetime.strptime(display_slot.split("-")[0], "%H:%M").time()
                    end_time = datetime.strptime(display_slot.split("-")[1], "%H:%M").time()
                    lunch_start = lunch_schedule[dept]["start"]
                    lunch_end = lunch_schedule[dept]["end"]
                    if (start_time >= morning_break_start and start_time < morning_break_end) or \
                       (end_time > morning_break_start and end_time <= morning_break_end):
                        html_content += '<td class="break-cell">Morning Break</td>'
                        continue
                    if (start_time >= lunch_start and start_time < lunch_end) or \
                       (end_time > lunch_start and end_time <= lunch_end):
                        dept_lunch = lunch_schedule[dept]["start"].strftime("%H:%M") + "-" + lunch_schedule[dept]["end"].strftime("%H:%M")
                        html_content += f'<td class="break-cell">Lunch Break ({dept_lunch})</td>'
                        continue
                    slots = slot_mapping[display_slot]
                    cell_content = ""
                    cell_style = ""
                    current_course = None
                    for slot in slots:
                        info = schedule[day][slot].get(timetable_key, {})
                        if not info:
                            continue
                        if current_course and info["course_code"] != current_course:
                            break
                        current_course = info["course_code"]
                        cell_content = info["label"]
                        color = assign_color(info["course_code"])
                        cell_style = f'background-color: #{color};'
                    html_content += f'<td style="{cell_style}">{cell_content}</td>'
                html_content += '</tr>'
            html_content += '</tbody></table>'

# Add elective details table
html_content += '<h2 class="text-2xl font-semibold mt-8 mb-4">Elective Scheduling Details</h2>'
html_content += '<table class="elective-table">'
html_content += '<thead><tr><th>Basket ID</th><th>Course Name</th><th>Faculty</th><th>Room</th><th>Time Slot</th><th>LTPSC</th><th>Extra Sessions</th></tr></thead><tbody>'
for detail in elective_details:
    html_content += '<tr>'
    html_content += f'<td>{detail["Basket ID"]}</td>'
    html_content += f'<td>{detail["Course Name"]}</td>'
    html_content += f'<td>{detail["Faculty"]}</td>'
    html_content += f'<td>{detail["Room"]}</td>'
    html_content += f'<td>{detail["Time Slot"]}</td>'
    html_content += f'<td>{detail["LTPSC"]}</td>'
    html_content += f'<td>{detail["Extra Sessions"]}</td>'
    html_content += '</tr>'
html_content += '</tbody></table>'

html_content += """
</body>
</html>
"""

with open(os.path.join(output_dir, "timetable.html"), "w") as f:
    f.write(html_content)

# Generate Excel
logging.info("Generating Excel output")
wb = Workbook()
wb.remove(wb.active)

for dept, semesters in semesters_by_dept.items():
    for semester in semesters:
        semester_courses = courses_df[(courses_df["department"] == dept) & (courses_df["semester"] == semester)]
        for section_id in semester_courses["section_id"].unique():
            timetable_key = f"{dept}_{semester}_{section_id}"
            section = sections_df[sections_df["section_id"] == section_id].iloc[0]
            sheet_name = f"{section['batch_name']}_{semester}".replace("/", "_")
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append([""] * 3)
            ws.append(["INDIAN INSTITUTE OF INFORMATION TECHNOLOGY, DHARWAD"])
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(display_slots) + 1)
            ws.append(["Time Table for an Academic year Dec 24 – April 2025"])
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(display_slots) + 1)
            roll_start = f"{section['year']}{section['department'].lower()}{section['batch_name'][-2:]}001"
            roll_end = f"{section['year']}{section['department'].lower()}{section['batch_name'][-2:]}0{int(section['strength']):02d}"
            ws.append([f"Section: {section['batch_name']} – Roll no {roll_start} to {roll_end}"])
            ws.append([f"Group mail id – {section["year"]}{section["department"].lower()}{section["batch_name"][-2:]}@iiitdwd.ac.in"])
            ws.append(["Day"] + display_slots)
            
            for day in scheduling_days:
                row = [day]
                for display_slot in display_slots:
                    start_time = datetime.strptime(display_slot.split("-")[0], "%H:%M").time()
                    end_time = datetime.strptime(display_slot.split("-")[1], "%H:%M").time()
                    lunch_start = lunch_schedule[dept]["start"]
                    lunch_end = lunch_schedule[dept]["end"]
                    if (start_time >= morning_break_start and start_time < morning_break_end) or \
                       (end_time > morning_break_start and end_time <= morning_break_end):
                        row.append("Morning Break")
                        continue
                    if (start_time >= lunch_start and start_time < lunch_end) or \
                       (end_time > lunch_start and end_time <= lunch_end):
                        dept_lunch = lunch_schedule[dept]["start"].strftime("%H:%M") + "-" + lunch_schedule[dept]["end"].strftime("%H:%M")
                        row.append(f"Lunch Break ({dept_lunch})")
                        continue
                    slots = slot_mapping[display_slot]
                    cell_content = ""
                    current_course = None
                    for slot in slots:
                        info = schedule[day][slot].get(timetable_key, {})
                        if not info:
                            continue
                        if current_course and info["course_code"] != current_course:
                            break
                        current_course = info["course_code"]
                        cell_content = info["label"]
                    row.append(cell_content)
                ws.append(row)
            
            for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[9]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            
            for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                    if cell.column == 1:
                        continue
                    if cell.value in [None, ""]:
                        continue
                    if "Break" in str(cell.value):
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    else:
                        for course_code in course_colors:
                            if course_code in str(cell.value):
                                color = course_colors[course_code]
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                break
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

# Add elective details sheet
ws = wb.create_sheet(title="Elective_Details")
ws.append(["Basket ID", "Course Name", "Faculty", "Room", "Time Slot", "LTPSC", "Extra Sessions"])
for detail in elective_details:
    ws.append([
        detail["Basket ID"],
        detail["Course Name"],
        detail["Faculty"],
        detail["Room"],
        detail["Time Slot"],
        detail["LTPSC"],
        detail["Extra Sessions"]
    ])

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width

wb.save(os.path.join(output_dir, "timetable.xlsx"))

logging.info("Timetable generated successfully in the 'output' directory.")
print("Timetable generated successfully in the 'output' directory.")